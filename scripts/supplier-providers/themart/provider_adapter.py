#!/usr/bin/env python3
"""Repository-safe runtime paths and snapshot normalization for The Mart."""

from __future__ import annotations

import argparse
import asyncio
import csv
import importlib.util
import json
import os
from pathlib import Path
import re
import shutil
import sys
from typing import Mapping, NamedTuple


FRESHNESS_STATES = {"CURRENT_SNAPSHOT", "STALE", "UNKNOWN"}
DANGEROUS_SPREADSHEET_PREFIXES = ("=", "+", "-", "@", "\t", "\r")
STAGING_DIRECTORY = ".themart_capture_staging"


class ProviderConfigurationError(ValueError):
    """Raised when local-only provider configuration crosses repository boundaries."""


class UnsafeSpreadsheetError(RuntimeError):
    """Raised when a candidate supported spreadsheet contains an unsafe cell."""


class SupportedOutputError(RuntimeError):
    """Raised when capture output cannot cross the hardened publication boundary."""


class RuntimePaths(NamedTuple):
    browser_profile: Path
    output_root: Path


class SupportedCapture(NamedTuple):
    capture_dir: Path
    products_csv: Path
    products_xlsx: Path
    diagnostics_csv: Path
    snapshot_jsonl: Path
    data_zip: Path


def _is_within(path: Path, root: Path) -> bool:
    try:
        path.relative_to(root)
        return True
    except ValueError:
        return False


def _external_path(value: str, *, variable: str, repository_root: Path) -> Path:
    candidate = Path(value).expanduser()
    if not candidate.is_absolute():
        raise ProviderConfigurationError(f"{variable} must be an absolute local path")
    resolved = candidate.resolve()
    if _is_within(resolved, repository_root):
        raise ProviderConfigurationError(f"{variable} must remain outside the repository")
    return resolved


def resolve_runtime_paths(repository_root: Path | str, environ: Mapping[str, str]) -> RuntimePaths:
    """Resolve local-only runtime paths without creating files or starting a browser."""

    root = Path(repository_root).resolve()
    profile_value = environ.get("THEMART_BROWSER_PROFILE_DIR", "").strip()
    if not profile_value:
        raise ProviderConfigurationError("THEMART_BROWSER_PROFILE_DIR is required for live capture")
    profile = _external_path(
        profile_value,
        variable="THEMART_BROWSER_PROFILE_DIR",
        repository_root=root,
    )

    output_value = environ.get("THEMART_OUTPUT_DIR", "").strip()
    if output_value:
        output = _external_path(
            output_value,
            variable="THEMART_OUTPUT_DIR",
            repository_root=root,
        )
    else:
        output = (profile.parent / "themart_capture_output").resolve()
    return RuntimePaths(browser_profile=profile, output_root=output)


def configure_capture_module(capture_module: object, paths: RuntimePaths) -> None:
    """Apply validated local paths to the untouched exact-source capture module."""

    setattr(capture_module, "USER_DATA_DIR", paths.browser_profile)
    setattr(capture_module, "OUT_ROOT", paths.output_root / STAGING_DIRECTORY)


def prepare_capture_module(repository_root: Path | str, environ: Mapping[str, str]):
    """Load and configure the exact source without starting browser or network activity."""

    paths = resolve_runtime_paths(repository_root, environ)
    exact_source = Path(__file__).with_name("themart_capture.py")
    spec = importlib.util.spec_from_file_location("themart_exact_capture", exact_source)
    if spec is None or spec.loader is None:
        raise ProviderConfigurationError("cannot load exact The Mart capture source")
    capture_module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(capture_module)
    configure_capture_module(capture_module, paths)
    return capture_module, paths


def _load_exact_recovery_module():
    exact_source = Path(__file__).with_name("themart_extract_existing_html.py")
    spec = importlib.util.spec_from_file_location("themart_exact_recovery", exact_source)
    if spec is None or spec.loader is None:
        raise SupportedOutputError("cannot load exact The Mart recovery source")
    recovery_module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(recovery_module)
    return recovery_module


def _unsafe_spreadsheet_value(value: object) -> bool:
    return isinstance(value, str) and value.startswith(DANGEROUS_SPREADSHEET_PREFIXES)


def verify_supported_spreadsheets(paths: list[Path] | tuple[Path, ...]) -> None:
    """Fail closed if any supported CSV/XLSX cell can be interpreted as a formula."""

    from openpyxl import load_workbook

    for path in paths:
        candidate = Path(path)
        if not candidate.is_file():
            raise UnsafeSpreadsheetError(f"supported spreadsheet is missing: {candidate.name}")
        if candidate.suffix.lower() == ".csv":
            with candidate.open("r", encoding="utf-8-sig", newline="") as handle:
                for row_number, row in enumerate(csv.reader(handle), start=1):
                    for column_number, value in enumerate(row, start=1):
                        if _unsafe_spreadsheet_value(value):
                            raise UnsafeSpreadsheetError(
                                f"unsafe spreadsheet cell: {candidate.name}:{row_number}:{column_number}"
                            )
        elif candidate.suffix.lower() == ".xlsx":
            workbook = load_workbook(candidate, read_only=True, data_only=False)
            try:
                for sheet in workbook.worksheets:
                    for row in sheet.iter_rows():
                        for cell in row:
                            if _unsafe_spreadsheet_value(cell.value):
                                raise UnsafeSpreadsheetError(
                                    f"unsafe spreadsheet cell: {candidate.name}:{sheet.title}:{cell.coordinate}"
                                )
            finally:
                workbook.close()
        else:
            raise UnsafeSpreadsheetError(f"unsupported spreadsheet type: {candidate.name}")


def run_supported_capture(capture_module: object, paths: RuntimePaths) -> SupportedCapture:
    """Capture into private staging and publish only verified hardened outputs."""

    configure_capture_module(capture_module, paths)
    staging_root = paths.output_root / STAGING_DIRECTORY
    if staging_root.exists() and any(staging_root.iterdir()):
        raise SupportedOutputError("capture staging is not empty; refusing ambiguous publication")
    staging_root.mkdir(parents=True, exist_ok=True)

    asyncio.run(capture_module.main())
    capture_dirs = sorted(
        path for path in staging_root.glob("themart_capture_*") if path.is_dir()
    )
    if len(capture_dirs) != 1:
        raise SupportedOutputError(
            f"expected exactly one staged capture directory, found {len(capture_dirs)}"
        )
    staged_capture = capture_dirs[0]

    recovery = _load_exact_recovery_module()
    try:
        recovery.main(["--strict", str(staged_capture)])
    except SystemExit as exc:
        if exc.code not in (None, 0):
            raise SupportedOutputError(f"hardened recovery failed with exit code {exc.code}") from exc

    staged_indexes = staged_capture / "indexes"
    products_csv = staged_indexes / "products_raw_recovered.csv"
    products_xlsx = staged_indexes / "products_raw_recovered.xlsx"
    diagnostics_csv = staged_indexes / "extraction_diagnostics.csv"
    verify_supported_spreadsheets((products_csv, products_xlsx, diagnostics_csv))

    snapshot_jsonl = staged_indexes / "supplier_snapshot.jsonl"
    normalize_recovered_csv(
        products_csv,
        snapshot_jsonl,
        freshness_state="CURRENT_SNAPSHOT",
    )

    raw_indexes = (
        staged_indexes / "products_raw.csv",
        staged_indexes / "products_raw.xlsx",
        staged_indexes / "captured_pages.csv",
    )
    for raw_path in raw_indexes:
        raw_path.unlink(missing_ok=True)
    (staged_capture / "capture_summary.txt").unlink(missing_ok=True)
    (staging_root / f"{staged_capture.name}.zip").unlink(missing_ok=True)

    staged_data_zip = staging_root / f"{staged_capture.name}_recovered_indexes_only.zip"
    if not staged_data_zip.is_file():
        raise SupportedOutputError("hardened data-only ZIP is missing")

    paths.output_root.mkdir(parents=True, exist_ok=True)
    published_capture = paths.output_root / staged_capture.name
    published_zip = paths.output_root / staged_data_zip.name
    if published_capture.exists() or published_zip.exists():
        raise SupportedOutputError("supported output destination already exists")
    shutil.move(str(staged_capture), str(published_capture))
    shutil.move(str(staged_data_zip), str(published_zip))
    if staging_root.exists() and not any(staging_root.iterdir()):
        staging_root.rmdir()

    published_indexes = published_capture / "indexes"
    return SupportedCapture(
        capture_dir=published_capture,
        products_csv=published_indexes / products_csv.name,
        products_xlsx=published_indexes / products_xlsx.name,
        diagnostics_csv=published_indexes / diagnostics_csv.name,
        snapshot_jsonl=published_indexes / snapshot_jsonl.name,
        data_zip=published_zip,
    )


def _money(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("€", "").replace(" ", "")
    if not text:
        return None
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _vat_rate(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        numeric = float(value)
    else:
        match = re.search(r"\d+(?:[.,]\d+)?", str(value))
        if not match:
            return None
        numeric = float(match.group(0).replace(",", "."))
    return numeric / 100 if numeric > 1 else numeric


def _integer(value: object) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _sanitized_filename(value: object) -> str | None:
    if value is None or value == "":
        return None
    return Path(str(value)).name or None


def normalize_recovered_row(row: Mapping[str, object], *, freshness_state: str) -> dict:
    """Map one recovered evidence row to supplier_price_snapshot.schema.json."""

    if freshness_state not in FRESHNESS_STATES:
        raise ValueError(f"unsupported freshness_state: {freshness_state}")

    product_name = str(row.get("product_name") or "").strip()
    captured_at = str(row.get("captured_at") or "").strip()
    if not product_name:
        raise ValueError("product_name is required; do not invent supplier identity")
    if not captured_at:
        raise ValueError("captured_at is required; do not invent snapshot freshness")

    warnings: list[str] = []
    pack = str(row.get("pack_description") or row.get("pack") or "").strip() or None
    unit = str(row.get("unit") or "").strip()
    if pack is None:
        warnings.append("pack_missing")
    if not unit:
        unit = "UNKNOWN"
        warnings.append("unit_missing")

    price_net = _money(row.get("product_price_without_vat"))
    price_gross = _money(row.get("product_price_with_vat"))
    unit_net = _money(row.get("unit_price_without_vat"))
    unit_gross = _money(row.get("unit_price_with_vat"))
    if price_net is None and price_gross is None and unit_net is None and unit_gross is None:
        warnings.append("price_fields_missing")

    vat_rate = _vat_rate(row.get("vat_percent_guess") or row.get("vat_rate"))
    vat_status = "INFERRED" if vat_rate is not None else "UNKNOWN"
    if vat_status == "UNKNOWN":
        warnings.append("vat_status_unknown")

    if unit_net is not None:
        unit_price = unit_net
        unit_price_basis = "NET"
    elif unit_gross is not None:
        unit_price = unit_gross
        unit_price_basis = "GROSS"
    else:
        unit_price = None
        unit_price_basis = "UNKNOWN"

    raw_reference = _sanitized_filename(row.get("source_html_file"))
    product_url = str(row.get("product_url") or "").strip() or None
    category_url = str(row.get("category_url") or "").strip() or None
    source_reference = product_url or category_url or raw_reference
    if not source_reference:
        source_reference = "UNKNOWN"
        warnings.append("source_reference_missing")

    confidence = "HIGH" if not warnings else ("MEDIUM" if len(warnings) <= 2 else "LOW")
    evidence_state = "NORMALIZED" if not warnings else "NEEDS_REVIEW"

    return {
        "provider": "themart",
        "captured_at": captured_at,
        "root_category": str(row.get("root_category") or "").strip() or None,
        "category_name": str(row.get("category_name") or "").strip() or None,
        "page_num": _integer(row.get("page_num")),
        "sku": str(row.get("sku") or "").strip() or None,
        "product_code": str(row.get("product_code") or "").strip() or None,
        "product_name": product_name,
        "pack": pack,
        "unit": unit,
        "price_net": price_net,
        "price_gross": price_gross,
        "vat_rate": vat_rate,
        "vat_status": vat_status,
        "unit_price": unit_price,
        "unit_price_basis": unit_price_basis,
        "source_reference": source_reference,
        "product_url": product_url,
        "category_url": category_url,
        "evidence_state": evidence_state,
        "confidence": confidence,
        "freshness_state": freshness_state,
        "parsing_warnings": warnings,
        "raw_capture_reference": raw_reference,
    }


def normalize_recovered_csv(
    source: Path | str,
    destination: Path | str,
    *,
    freshness_state: str,
) -> int:
    """Write recovered CSV evidence as deterministic canonical JSON Lines."""

    source_path = Path(source)
    destination_path = Path(destination)
    with source_path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))

    records = [normalize_recovered_row(row, freshness_state=freshness_state) for row in rows]
    destination_path.parent.mkdir(parents=True, exist_ok=True)
    with destination_path.open("w", encoding="utf-8", newline="\n") as handle:
        for record in records:
            handle.write(json.dumps(record, ensure_ascii=False, sort_keys=True, separators=(",", ":")))
            handle.write("\n")
    return len(records)


def main(argv: list[str] | None = None, *, environ: Mapping[str, str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    commands = parser.add_subparsers(dest="command", required=True)

    capture_parser = commands.add_parser(
        "capture",
        help="Run an explicit foreground capture with manual browser login.",
    )
    capture_parser.add_argument(
        "--prepare-only",
        action="store_true",
        help="Validate local-only configuration without creating state or launching a browser.",
    )

    normalize_parser = commands.add_parser(
        "normalize",
        help="Normalize a recovered CSV to supplier snapshot JSON Lines.",
    )
    normalize_parser.add_argument("source_csv")
    normalize_parser.add_argument("destination_jsonl")
    normalize_parser.add_argument(
        "--freshness-state",
        choices=sorted(FRESHNESS_STATES),
        required=True,
    )

    args = parser.parse_args(argv)
    environment = os.environ if environ is None else environ
    repository_root = Path(__file__).resolve().parents[3]

    if args.command == "normalize":
        count = normalize_recovered_csv(
            args.source_csv,
            args.destination_jsonl,
            freshness_state=args.freshness_state,
        )
        print(f"Normalized supplier evidence rows: {count}")
        return 0

    try:
        capture_module, paths = prepare_capture_module(repository_root, environment)
    except ProviderConfigurationError as exc:
        print(f"The Mart live capture configuration: FAIL ({exc})", file=sys.stderr)
        return 2
    if args.prepare_only:
        print("The Mart live capture configuration: READY")
        return 0

    result = run_supported_capture(capture_module, paths)
    print(f"Supported hardened capture: {result.capture_dir.name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
