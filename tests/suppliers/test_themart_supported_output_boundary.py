from __future__ import annotations

import csv
from io import BytesIO, StringIO
import importlib.util
import json
from pathlib import Path
from types import SimpleNamespace
import zipfile

from openpyxl import Workbook, load_workbook
import pytest
import yaml


ROOT = Path(__file__).resolve().parents[2]
PROVIDER = ROOT / "scripts/supplier-providers/themart"
ADAPTER = PROVIDER / "provider_adapter.py"
DEPENDENCY_EVIDENCE = PROVIDER / "dependency_evidence.yaml"
PROVIDER_CONTRACT = PROVIDER / "provider_contract.yaml"
DANGEROUS_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _load_adapter():
    spec = importlib.util.spec_from_file_location("themart_supported_boundary", ADAPTER)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _write_unsafe_spreadsheets(indexes: Path) -> None:
    indexes.mkdir(parents=True, exist_ok=True)
    with (indexes / "products_raw.csv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["product_name"])
        writer.writerow(["=UNSAFE()"])
    workbook = Workbook()
    workbook.active["A1"] = "=UNSAFE()"
    workbook.save(indexes / "products_raw.xlsx")
    (indexes / "captured_pages.csv").write_text("title\n=UNSAFE()\n", encoding="utf-8")


class FakeExactCapture:
    USER_DATA_DIR: Path | None = None
    OUT_ROOT: Path | None = None

    async def main(self) -> None:
        assert self.OUT_ROOT is not None
        capture = self.OUT_ROOT / "themart_capture_20260904_120000"
        pages = capture / "pages_html"
        pages.mkdir(parents=True)
        (pages / "pantry__test__page_001.html").write_text(
            """
            <html><body><div class="product-card">
              <span>=UNSAFE</span><span>11111</span>
              <span>Τιμή / τεμάχιο χωρίς ΦΠΑ</span><span>1,00 €</span>
            </div></body></html>
            """,
            encoding="utf-8",
        )
        _write_unsafe_spreadsheets(capture / "indexes")
        (capture / "capture_summary.txt").write_text("historical raw export", encoding="utf-8")
        with zipfile.ZipFile(self.OUT_ROOT / f"{capture.name}.zip", "w") as archive:
            archive.write(capture / "indexes/products_raw.csv", "indexes/products_raw.csv")


def test_supported_capture_publishes_only_hardened_spreadsheets_and_removes_raw_bypass(tmp_path):
    adapter = _load_adapter()
    paths = adapter.RuntimePaths(
        browser_profile=tmp_path / "profile",
        output_root=tmp_path / "supported-output",
    )
    capture = FakeExactCapture()
    adapter.configure_capture_module(capture, paths)

    result = adapter.run_supported_capture(capture, paths)

    assert result.capture_dir == paths.output_root / "themart_capture_20260904_120000"
    assert result.products_csv == result.capture_dir / "indexes/products_raw_recovered.csv"
    assert result.products_xlsx == result.capture_dir / "indexes/products_raw_recovered.xlsx"
    assert result.snapshot_jsonl == result.capture_dir / "indexes/supplier_snapshot.jsonl"
    assert result.data_zip == paths.output_root / "themart_capture_20260904_120000_recovered_indexes_only.zip"
    assert result.products_csv.is_file()
    assert result.products_xlsx.is_file()
    assert result.snapshot_jsonl.is_file()
    assert result.data_zip.is_file()

    with zipfile.ZipFile(result.data_zip) as archive:
        archive_names = set(archive.namelist())
        assert "indexes/products_raw.csv" not in archive_names
        assert "indexes/products_raw.xlsx" not in archive_names
        assert "indexes/captured_pages.csv" not in archive_names
        archived_csv = archive.read("indexes/products_raw_recovered.csv").decode("utf-8-sig")
        archived_row = next(csv.DictReader(StringIO(archived_csv)))
        assert archived_row["product_name"] == "'=UNSAFE"
        archived_workbook = load_workbook(BytesIO(archive.read("indexes/products_raw_recovered.xlsx")), data_only=False)
        archived_values = [
            cell.value
            for row in archived_workbook.active.iter_rows()
            for cell in row
            if isinstance(cell.value, str)
        ]
        assert not any(value.startswith(DANGEROUS_PREFIXES) for value in archived_values)

    indexes = result.capture_dir / "indexes"
    assert not (indexes / "products_raw.csv").exists()
    assert not (indexes / "products_raw.xlsx").exists()
    assert not (indexes / "captured_pages.csv").exists()
    assert not (paths.output_root / "themart_capture_20260904_120000.zip").exists()
    assert not (paths.output_root / ".themart_capture_staging").exists()

    with result.products_csv.open("r", encoding="utf-8-sig", newline="") as handle:
        row = next(csv.DictReader(handle))
    assert row["product_name"] == "'=UNSAFE"

    workbook = load_workbook(result.products_xlsx, data_only=False)
    values = [cell.value for row in workbook.active.iter_rows() for cell in row if isinstance(cell.value, str)]
    assert "'=UNSAFE" in values
    assert not any(value.startswith(DANGEROUS_PREFIXES) for value in values)


@pytest.mark.parametrize("prefix", DANGEROUS_PREFIXES)
def test_supported_spreadsheet_gate_rejects_every_dangerous_prefix(tmp_path, prefix):
    adapter = _load_adapter()
    csv_path = tmp_path / "unsafe.csv"
    xlsx_path = tmp_path / "unsafe.xlsx"
    with csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["value"])
        writer.writerow([prefix + "payload"])
    workbook = Workbook()
    workbook.active["A1"] = prefix + "payload"
    workbook.save(xlsx_path)

    with pytest.raises(adapter.UnsafeSpreadsheetError):
        adapter.verify_supported_spreadsheets([csv_path, xlsx_path])


@pytest.mark.parametrize("prefix", DANGEROUS_PREFIXES)
def test_supported_capture_fails_closed_before_publication_for_every_dangerous_prefix(
    tmp_path, monkeypatch, prefix
):
    adapter = _load_adapter()
    paths = adapter.RuntimePaths(tmp_path / "profile", tmp_path / "supported-output")
    capture = FakeExactCapture()

    class UnsafeRecovery:
        @staticmethod
        def main(argv):
            capture_dir = Path(argv[-1])
            indexes = capture_dir / "indexes"
            with (indexes / "products_raw_recovered.csv").open("w", encoding="utf-8", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerow(["product_name"])
                writer.writerow([prefix + "payload"])
            workbook = Workbook()
            workbook.active["A1"] = prefix + "payload"
            workbook.save(indexes / "products_raw_recovered.xlsx")
            (indexes / "extraction_diagnostics.csv").write_text("status\nok\n", encoding="utf-8")
            zip_path = capture_dir.parent / f"{capture_dir.name}_recovered_indexes_only.zip"
            with zipfile.ZipFile(zip_path, "w") as archive:
                archive.write(indexes / "products_raw_recovered.csv", "indexes/products_raw_recovered.csv")

    monkeypatch.setattr(adapter, "_load_exact_recovery_module", lambda: UnsafeRecovery())

    with pytest.raises(adapter.UnsafeSpreadsheetError):
        adapter.run_supported_capture(capture, paths)

    capture_name = "themart_capture_20260904_120000"
    assert not (paths.output_root / capture_name).exists()
    assert not (paths.output_root / f"{capture_name}_recovered_indexes_only.zip").exists()
    assert (paths.output_root / ".themart_capture_staging" / capture_name).is_dir()


def test_capture_cli_dispatches_through_supported_boundary_without_direct_exact_bypass(tmp_path, monkeypatch):
    adapter = _load_adapter()
    marker = tmp_path / "supported-boundary-used"
    paths = adapter.RuntimePaths(tmp_path / "profile", tmp_path / "output")

    class ExactCaptureMustNotRunDirectly:
        async def main(self):
            raise AssertionError("CLI bypassed supported output boundary")

    monkeypatch.setattr(adapter, "prepare_capture_module", lambda _root, _env: (ExactCaptureMustNotRunDirectly(), paths))

    def supported_executor(_capture, _paths):
        marker.write_text("used", encoding="utf-8")
        return SimpleNamespace(capture_dir=paths.output_root / "published-capture")

    monkeypatch.setattr(adapter, "run_supported_capture", supported_executor, raising=False)

    assert adapter.main(["capture"], environ={"THEMART_BROWSER_PROFILE_DIR": str(tmp_path / "profile")}) == 0
    assert marker.read_text(encoding="utf-8") == "used"


def test_provider_contract_exposes_only_the_fail_closed_hardened_capture_route():
    contract = yaml.safe_load(PROVIDER_CONTRACT.read_text(encoding="utf-8"))
    execution = contract["supported_execution"]

    assert execution["entrypoint"] == "scripts/supplier-providers/themart/provider_adapter.py capture"
    assert execution["direct_historical_scripts_supported"] is False
    assert execution["raw_spreadsheet_exports_supported"] is False
    assert execution["publication_mode"] == "fail_closed_after_recovery_and_formula_scan"
    assert execution["formula_prefixes_rejected"] == ["=", "+", "-", "@", "TAB", "CR"]
    assert execution["bypass_option_available"] is False


def test_playwright_pin_has_audited_local_compatibility_evidence_and_no_conflict():
    evidence = yaml.safe_load(DEPENDENCY_EVIDENCE.read_text(encoding="utf-8"))
    playwright = evidence["playwright"]

    assert playwright["version"] == "1.49.1"
    assert playwright["selection_rationale"].startswith("match the installed Playwright package")
    assert playwright["evidence_source"] == "owner-local themart_capture_tool virtual-environment package metadata"
    assert playwright["observed_python"] == "3.12.13"
    assert "exact source imports playwright.async_api" in playwright["compatibility_evidence"]
    assert playwright["relationship_to_original_bytes"] == "COMPANION_DEPENDENCY_NOT_PART_OF_EXACT_SOURCE"

    declarations = []
    for path in ROOT.rglob("*"):
        if not path.is_file() or ".git" in path.parts or path == Path(__file__):
            continue
        if path.suffix.lower() not in {".txt", ".toml", ".yaml", ".yml"}:
            continue
        for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
            stripped = line.strip().lower()
            if stripped.startswith("playwright") and any(token in stripped for token in ("==", ">=", "<=", "~=", "!=", ">", "<")):
                declarations.append((path.relative_to(ROOT).as_posix(), stripped))

    assert declarations == [("scripts/supplier-providers/themart/requirements-runtime.txt", "playwright==1.49.1")]
